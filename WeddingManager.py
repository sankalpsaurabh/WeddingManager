from configparser import ConfigParser
from msilib import Table
from msilib.schema import tables
from operator import truediv
from flask import Flask,render_template,request,redirect
from openpyxl import Workbook,load_workbook

app = Flask(__name__)

@app.route("/")
def homepage():
    return render_template("index.html")
    
@app.route("/GuestList",methods=['GET','POST'])
def GetGuestList():
    SheetNames = GetDataFromXL("")
    return render_template("GuestList.html",Tables=SheetNames)


@app.route("/GuestData/<string:SheetName>",methods=['GET','POST'])
def GetGuestData(SheetName):
    sheet = GetDataFromXL(SheetName)
    return render_template("GuestData.html", sheet=sheet,SheetName=SheetName)
   
@app.route("/EditSheet",methods=['GET','POST'])
def EditSheet():
    print("Editing Sheet")
    
    member = request.form['membername']
    group = request.form['group']
    gender = request.form['gender']
    plusmember = request.form['plusmember']
    expected = request.form['excepted']
    arrivaldate = request.form['arrivaldate']
    rowno = request.form['row']
    sheetname = request.form['sheet']
    url = "/GuestData/" + sheetname

    res = updatexl(sheetname,rowno,member,group,gender,plusmember,expected,arrivaldate)
    return redirect(url)

def updatexl(sheetname,rowno,member,group,gender,plusmember,expected,arrivaldate):
    wb = load_workbook("./Lib\\WeddingList.xlsx",data_only=True)
    row = int(rowno)
    worksheet = wb[sheetname]
    worksheet.cell(row = row,column= 1).value = member
    worksheet.cell(row = row,column= 2).value = group
    worksheet.cell(row = row,column= 3).value = gender
    worksheet.cell(row = row,column= 4).value = plusmember
    worksheet.cell(row = row,column= 5).value = expected
    worksheet.cell(row = row,column= 6).value = arrivaldate
    wb.save("./Lib\\WeddingList.xlsx")
    wb.close()


    return True

def GetDataFromXL(SheetName):
    try:
        wb = load_workbook("./Lib\\WeddingList.xlsx",data_only=True)
        if(SheetName == ""):
            wb.close()
            return wb.sheetnames
        else:
            print(SheetName)
            sheet = wb[SheetName]
            wb.close()
            return sheet
    except:
        ErrorHandler("GetDataFromXml")

def GetSheetData(worksheetname,worksheet):
    print("ok")

def ErrorHandler(Exception):
    return render_template("Error.html")
    
    
   


if __name__ == "__main__":
    app.run(debug=False, host="0.0.0.0",port=8000)
